import openpyxl


def add_to_excel(file_name, sheet_name, data, columns=None):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    if columns:
        for i, col_name in enumerate(columns, 1):
            sheet.cell(row=1, column=i, value=col_name)
        start_row = 2
    else:
        start_row = 1

    for i in range(len(data[0])):
        for j in range(len(data)):
            sheet.cell(row=start_row + i, column=j + 1, value=data[j][i])

    workbook.save(file_name)


# Example usage
data = [['Search', 'News', 'Blog', 'Recently viewed products', 'Compare products list', 'New products'],
        ['Sitemap', 'Shipping & returns', 'Privacy notice', 'Conditions of Use', 'About us', 'Contact us'],
        ['My account', 'Orders', 'Addresses', 'Shopping cart', 'Wishlist', 'Apply for vendor account']]
columns = ["Information", "Customer service", "My account"]
add_to_excel("example1.xlsx", "Sheet1", data, columns)
