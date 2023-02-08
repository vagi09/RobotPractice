from openpyxl import Workbook
from contextlib import closing
from openpyxl.styles import Font, Border, Side
import openpyxl


def make_excel_file(file_name):
    with closing(Workbook()) as wb:
        wb.save(file_name)


def add_to_excel(file_name, sheet_name, data, columns=None):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name

    bold_font = Font(bold=True, color="FF0000")
    thin_side = Side(style='thin')
    border = Border(top=thin_side, bottom=thin_side, left=thin_side, right=thin_side)

    if columns:
        for i, col_name in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=i, value=col_name)
            cell.font = bold_font
            cell.border = border
        start_row = 2
    else:
        start_row = 1

    for i in range(len(data[0])):
        for j in range(len(data)):
            cell = sheet.cell(row=start_row + i, column=j + 1, value=data[j][i])
            cell.border = border

    workbook.save(file_name)


